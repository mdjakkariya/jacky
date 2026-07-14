"""Resolve ``@path`` mentions in a user's message into bounded file context for the model.

The user points at a file with ``@path`` (any type). This reads/extracts it **on-device**
and prepends a bounded preamble so the model has the content without spending an extra
``read_file`` turn — while respecting context/cost by capping each file (a large file is
excerpted with a pointer to read the rest). Text/code/config files inline directly;
``pdf``/``docx``/``xlsx`` are extracted to text via lazy optional deps (the ``docs`` extra);
images and other binaries get a short note (a local text model can't view an image — the
user would switch to a vision-capable model). Never raises: any failure degrades to a note.
"""

from __future__ import annotations

import re
from pathlib import Path

_MENTION_CAP = 10_000  # max chars of a single attached file handed to the model inline
_MAX_MENTIONS = 10  # cap attachments per message so a stray @ can't blow up the context

# ``@token`` at line start or after whitespace (so ``user@host`` is never mistaken for one).
_MENTION_RE = re.compile(r"(?:^|\s)@(\S+)")

_IMAGE_EXTS = frozenset(
    {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg", ".bmp", ".ico", ".tiff", ".heic"}
)


def find_mentions(text: str) -> list[str]:
    """The unique ``@path`` tokens in ``text``, in order (trailing punctuation stripped)."""
    seen: set[str] = set()
    out: list[str] = []
    for match in _MENTION_RE.finditer(text):
        token = match.group(1).rstrip(".,;:!?)]}\"'")
        if token and token not in seen:
            seen.add(token)
            out.append(token)
    return out


def resolve_mentions(text: str, cwd: str, *, cap: int = _MENTION_CAP) -> str:
    """Return ``text`` with a bounded preamble of any ``@path`` files' content prepended.

    Unchanged when there are no mentions. Each file is resolved relative to ``cwd`` (``~`` and
    absolute paths honored), extracted to text, and capped at ``cap`` chars.
    """
    tokens = find_mentions(text)
    if not tokens:
        return text
    blocks: list[str] = []
    for token in tokens[:_MAX_MENTIONS]:
        path = (Path(cwd) / token).expanduser()
        blocks.append(f"@{token}\n{extract_file(path, cap=cap)}")
    if len(tokens) > _MAX_MENTIONS:
        blocks.append(f"(+{len(tokens) - _MAX_MENTIONS} more attachments omitted)")
    preamble = "[Files the user attached with @ — use them as context]\n\n" + "\n\n".join(blocks)
    return f"{preamble}\n\n---\n\n{text}"


def extract_file(path: Path, *, cap: int = _MENTION_CAP) -> str:
    """A bounded text representation of ``path`` for the model (never raises)."""
    if not path.exists():
        return "(file not found)"
    if not path.is_file():
        return "(not a file)"
    try:
        size = _human_size(path.stat().st_size)
    except OSError:
        size = "?"
    suffix = path.suffix.lower()
    if suffix in _IMAGE_EXTS:
        raw = (
            f"(image, {size} — a local text model can't view images; switch to a "
            f"vision-capable model to analyze it)"
        )
    elif suffix == ".pdf":
        raw = _extract_pdf(path, size)
    elif suffix == ".docx":
        raw = _extract_docx(path, size)
    elif suffix == ".xlsx":
        raw = _extract_xlsx(path, size)
    else:
        raw = _read_text(path, size)
    return _bound(raw, cap)


def _read_text(path: Path, size: str) -> str:
    """Read a text file as UTF-8; a decode failure means it's binary → a note."""
    try:
        return path.read_text(encoding="utf-8")
    except (UnicodeDecodeError, ValueError):
        return f"(binary file, {size} — not shown)"
    except OSError as exc:
        return f"(couldn't read file: {exc})"


def _extract_pdf(path: Path, size: str) -> str:
    """Extract PDF text on-device via pypdf (lazy; a note if the extra isn't installed)."""
    try:
        from pypdf import PdfReader
    except ImportError:
        return f"(pdf, {size} — install the 'docs' extra to extract its text: uv sync --extra docs)"
    try:
        reader = PdfReader(str(path))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        return text.strip() or "(empty pdf)"
    except Exception as exc:  # a malformed PDF must not crash the turn
        return f"(couldn't read pdf: {exc})"


def _extract_docx(path: Path, size: str) -> str:
    """Extract .docx paragraph text on-device via python-docx (lazy)."""
    try:
        import docx
    except ImportError:
        return f"(doc, {size} — install the 'docs' extra to extract its text: uv sync --extra docs)"
    try:
        document = docx.Document(str(path))
        return "\n".join(p.text for p in document.paragraphs).strip() or "(empty document)"
    except Exception as exc:
        return f"(couldn't read doc: {exc})"


def _extract_xlsx(path: Path, size: str) -> str:
    """Extract .xlsx cells as tab-separated rows on-device via openpyxl (lazy)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return (
            f"(sheet, {size} — install the 'docs' extra to extract its cells: uv sync --extra docs)"
        )
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        lines: list[str] = []
        for ws in wb.worksheets:
            lines.append(f"# sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                lines.append("\t".join("" if c is None else str(c) for c in row))
        wb.close()
        return "\n".join(lines).strip() or "(empty sheet)"
    except Exception as exc:
        return f"(couldn't read sheet: {exc})"


def _bound(text: str, cap: int) -> str:
    """Cap extracted text to ~``cap`` chars (head + tail), pointing at read_file for the rest."""
    if len(text) <= cap:
        return text
    head = text[: cap * 35 // 100].rsplit("\n", 1)[0]
    tail = text[-(cap * 65 // 100) :].split("\n", 1)[-1]
    elided = len(text) - len(head) - len(tail)
    return (
        f"{head}\n\n[… {elided} chars elided — read the file with read_file/grep for the rest …]"
        f"\n\n{tail}"
    )


def _human_size(n: int) -> str:
    """Human-readable byte count (e.g. ``12.3 KB``)."""
    size = float(n)
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024 or unit == "GB":
            return f"{int(size)} {unit}" if unit == "B" else f"{size:.1f} {unit}"
        size /= 1024
    return f"{n} B"
